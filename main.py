#!/usr/bin/env python
# -*- coding: utf-8 -*-

import openpyxl
import json

def excel_to_json(excel_file, json_file):
    # Load the Excel workbook
    wb = openpyxl.load_workbook(excel_file)

    # Select the active worksheet
    ws = wb.active

    #Initialize an empty list to store the rows
    data = []

    # Iterate through each row in the worksheet
    for row in ws.iter_rows(values_only=True):
        # Append each row as a dictionary with column headers as keys
        data.append({ws.cell(row=1, column=col_idx).value: cell_value for col_idx, cell_value in enumerate(row, start=1)})

    # Close the workbook
    wb.close()

    # Write the data to a JSON file
    with open(json_file, 'w') as f:
        json.dump(data, f, indent=4)

def json_to_excel(json_file, excel_file):
    # Load JSON data
    with open(json_file, 'r') as f:
        data = json.load(f)

    # Create a new workbook
    wb = openpyxl.Workbook()

    # Select the active worksheet
    ws = wb.active

    # Write data from JSON to Excel
    for row_idx, row_data in enumerate(data, start=1):
        for col_idx, (header, value) in enumerate(row_data.items(), start=1):
            # Write header in the first row
            if row_idx == 1:
                ws.cell(row=row_idx, column=col_idx).value = header
            # Write data in subsequent rows
            ws.cell(row=row_idx + 1, column=col_idx).value = value

    # Save the workbook
    wb.save(excel_file)

    # Close the workbook
    wb.close()

def main():
    """ Main program """
    # Code goes over here.
    print("Main started")

    # Run function
    excel_to_json('Bok.xlsx', 'output.json')

    # Print
    print("The end.")

    # Convert json to xlsx
    json_to_excel('output.json', 'Bok_new.xlsx')

    return 0

if __name__ == "__main__":
    main()



